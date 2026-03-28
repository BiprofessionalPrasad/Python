
# -*- coding: utf-8 -*-
"""
Created on Mon Jul 27 11:47:58 2020

@author: pgaiton

Import excel.
Split by contents of first column.
Write into separate excels.
"""

import os
import pandas as pd

# ===================================================== # 
#                   Set Input Parameters                #
# ===================================================== # 
path=r'''K:\Projects\Finance\Jana\Budgeting\Export for Budget Revenue Plans_202005\outputs'''
os.chdir(path)
file='Output_20211214'; ext='.xlsx'; sheet='Sheet1'
excelfile=pd.ExcelFile(file+ext)
header_rows_to_drop=0;
bottom_rows_to_drop=0;
column_names_to_drop=['Unnamed: 0','Unnamed: 3']
group_by_column='Partner'
output_file_name='Python_'+file

# ----------- read excel into py data frame ------------------
dfExcel=pd.read_excel(excelfile,sheet,header=header_rows_to_drop,na_values=['NA'])

# --- drop blank column
#for col in column_names_to_drop:
#    dfExcel=dfExcel.drop(columns=col) 
# len(dfExcel)
# --- drop last row
# dfExcel.drop(dfExcel.tail(bottom_rows_to_drop).index,inplace=True) 
# -- not needed. write project partners into list
# ProjectPartners=dfExcel['Project Partner'].unique().tolist()
# -- transform dataframe column into a dictionary
#-- drop 1st column
dfExcel = dfExcel.drop('Index', 1)

dfs = dict(tuple(dfExcel.groupby(group_by_column)))

from openpyxl import load_workbook
from openpyxl import Workbook

# ---------------- write each key to separate tabs -------------
wb = Workbook()
wb.save(output_file_name+ext)
book = load_workbook(output_file_name+ext)
writer = pd.ExcelWriter(output_file_name+ext, engine='openpyxl') 
for df_name, df in dfs.items():
    # print(df_name)
    df.to_excel(writer,sheet_name=str(df_name),index=False)
writer.save()

'''
# ---------------- write each key to separate excel files -------------
for df_name, df in dfs.items():
    wb = Workbook()
    wb.save(df_name+ext)
    book = load_workbook(df_name+ext)
    writer = pd.ExcelWriter(df_name+ext, engine='openpyxl') 
    df.to_excel(writer, sheet_name=df_name,index=0)
    writer.save()
'''


